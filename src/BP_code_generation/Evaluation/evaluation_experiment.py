import time
from codebleu import calc_codebleu
from enum import Enum
import os
import pandas as pd
from openpyxl import load_workbook
import ast


#This file is in src/BP_code_generation/Evaluation/evaluation_expirement.py
#I need to import the MyOpenAIApi class from src/BP_code_generation/MyOpenAIApi.py


from src.BP_code_generation.myOpenAiApi import MyOpenAIApi

class BehaviorInstructionType(Enum):
    Basic = os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/1Basic Behavior Bot instructions"
    BasicPlus = os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/2BasicPlus Behavior Bot instructions"
    BasicPlusNoExamples = os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/2BasicPlus Behavior Bot instructions NoExamples"
    DSL = os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/3DSLs Behavior Bot instructions"
    DSL_Isolated= os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/3_5DSLs and isolation Behavior Bot instructions"
    Analysis = os.getcwd() + "/src/BP_code_generation/Instructions/All_Behavior_Instructions/4Analysis Behavior Bot instructions"
    Default = os.getcwd() + "/src/BP_code_generation/Instructions/Behavior Bot Instructions"




#first we need to create the needed context:
file_path = "C:/Users/Ron Ziskind/Desktop/thesis/Generating-Behavioral-Programs-Using-Large-Language-Models/src/BP_code_generation/Evaluation/Data/hot_cold_clean.js"

def get_data_set_of_file_and_eval(file_path_of_generated_code, output_file_path):
    #The function will return the data set of the file:
    # 1. The entities and queries part
    # 2. A dictionary of requirements and their corresponding code
    #we will read the file, then we will set the history accordingly
    with open(file_path_of_generated_code, "r") as file:
        lines = file.readlines()
    history = []
    #user messages are in comments/**/, assistant messages are in the code, both can be multi-lined
    user_message = ""
    assistant_message = ""
    isUser = False
    isBehavior = False
    requirements_and_code = {}
    for line in lines:
        if line.strip().startswith("//BEHAVIOR REQUIREMENTS"):
            isBehavior = True
            if assistant_message:
                history.append({"role": "assistant", "content": assistant_message})
                assistant_message = ""
                user_message = ""

            continue
        if line.strip().startswith("/*"):
            isUser = True
            if assistant_message:
                if not isBehavior:
                    history.append({"role": "assistant", "content": assistant_message})
                else:
                    requirements_and_code[user_message] = assistant_message
                assistant_message = ""
                user_message = ""
            # user_message = line
        elif line.strip().startswith("*/"):
            isUser = False
            if not isBehavior:
                history.append({"role": "user", "content": user_message})
            
        elif isUser:
            user_message += line
        else:
            assistant_message += line
    if assistant_message:
        if not isBehavior:
            history.append({"role": "assistant", "content": assistant_message})
        else:
            requirements_and_code[user_message] = assistant_message
    
    #history now holds the entity and queries part
    #now we will get the requirements and their corresponding code

    # print (requirements_and_code)
    # print (history)
    #export the history and requirements_and_code to a file called history_and_requirements
    with open("history_and_requirements.txt", "w") as file:
        file.write("History:\n")
        for item in history:
            file.write(f"Role: {item['role']}\n")
            file.write(f"Content: {item['content']}\n")
            file.write("\n")
        
        file.write("Requirements and Code:\n")
        for req, code in requirements_and_code.items():
            file.write(f"Requirement: {req}\n")
            file.write(f"Code: {code}\n")
            file.write("\n")

        
    
    
    # behavior_instructions_type = input("\n\n\Select behavior instructions type: \n1 for Basic, \n2 for BasicPlus, \n3 for DSL, \n3.5 for DSL_Isolated, \n4 for Analysis, \n5 for Default, \n6 for all \n see readme for more information\n your choice:")
    # behavior_instructions_types = []
    # if behavior_instructions_type == "1":
    #     behavior_instructions_types = [BehaviorInstructionType.Basic]
    # elif behavior_instructions_type == "2":
        behavior_instructions_types = [BehaviorInstructionType.BasicPlus]
    # elif behavior_instructions_type == "3":
    #     behavior_instructions_types = [BehaviorInstructionType.DSL]
    # elif behavior_instructions_type == "3.5":
    #     behavior_instructions_types = [BehaviorInstructionType.DSL_Isolated]
    # elif behavior_instructions_type == "4":
    #     behavior_instructions_types = [BehaviorInstructionType.Analysis]
    # elif behavior_instructions_type == "5":
    #     behavior_instructions_types = [BehaviorInstructionType.Default]
    # else:
    #     behavior_instructions_types = [BehaviorInstructionType.Basic, BehaviorInstructionType.BasicPlus, BehaviorInstructionType.DSL, BehaviorInstructionType.DSL_Isolated, BehaviorInstructionType.Analysis]
    behavior_instructions_types = [BehaviorInstructionType.BasicPlus, BehaviorInstructionType.BasicPlusNoExamples]
    behavior_instructions_types = [BehaviorInstructionType.BasicPlusNoExamples]
    
    
    all_models= []
    for behavior_instructions_type in behavior_instructions_types:
        model = MyOpenAIApi(model="gpt-4-turbo-2024-04-09", temp=0)
        behavior_instructions_file_path = behavior_instructions_type.value
        model.history = history
        
        #add first value to be system instructions
        instructions = ""
        with open(behavior_instructions_file_path, "r") as file:
                instructions = file.read()

        model.history.insert(0, {"role": "system", "content": instructions})
        model.History_For_Output = history.copy()
        all_models.append(model)

    requirements_code_generated_blue = {}
    file_name = file_path_of_generated_code.split("/")[-1].split("\\")[-1]
    for model in all_models:
        for req, code in requirements_and_code.items():

            # response = model.chat_with_gpt_cumulative(req)
            response = model.chat_with_gpt_cumulative_isolated_BP_generation(req)
            print("_____________________REQ_start______________________")
            print("_____________ChatGPT___________:\n", response)
            # print("_____________History___________:\n","\n".join([item["content"] for item in model.history[1:]]))
            print("_____________________REQ_end______________________")
            result = calc_codebleu([code], [response], lang="javascript", weights=(0.1, 0.1, 0.4, 0.4), tokenizer=None)
            requirements_code_generated_blue[req] = {"code": code, "generated_code": response, "blue_score": result, "file_name": file_name, "behavior_instructions_type": behavior_instructions_type.name}
            print(result)
    #export the requirements_code_generated_blue to a file called requirements_code_generated_blue
    with open("src/BP_code_generation/Evaluation/requirements_code_generated_blue.txt", "w") as file:
        for req, data in requirements_code_generated_blue.items():
            file.write(f"Requirement: {req}\n")
            file.write(f"Code: {data['code']}\n")
            file.write(f"Generated Code: {data['generated_code']}\n")
            file.write(f"Blue Score: {data['blue_score']}\n")
            file.write("\n")

    #export it to an excel file
    #in the next format:
    # | Requirement | Code | Generated Code | Blue Score | File Name | Behavior Instructions Type |
    # Convert the dictionary to a DataFrame
    df = pd.DataFrame.from_dict(requirements_code_generated_blue, orient='index')

    # Reset index to get 'Requirement' as a column
    df.reset_index(inplace=True)

    # Rename columns
    df.columns = ['Requirement', 'Code', 'Generated Code', 'Blue Score', 'Filename', 'Behavior Instructions Type']

    # Export to Excel
    if not output_file_path:
        output_file_path = "src/BP_code_generation/Evaluation/requirements_code_generated_blue.xlsx"
    #if the file already exists, we will append to it
    if os.path.exists(output_file_path):
        # Load the existing Excel file
        excel_file = output_file_path
        book = load_workbook(excel_file)

        # Load the existing sheet into a pandas DataFrame
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    else:
        df.to_excel(output_file_path, index=False)

    print("Exported to Excel file: ", output_file_path)




def get_data_set_of_file_and_eval_compare_isolated(file_path_of_generated_code, output_file_path):
    #The function will return the data set of the file:
    # 1. The entities and queries part
    # 2. A dictionary of requirements and their corresponding code
    #we will read the file, then we will set the history accordingly
    with open(file_path_of_generated_code, "r") as file:
        lines = file.readlines()
    history = []
    #user messages are in comments/**/, assistant messages are in the code, both can be multi-lined
    user_message = ""
    assistant_message = ""
    isUser = False
    isBehavior = False
    requirements_and_code = {}
    for line in lines:
        if line.strip().startswith("//BEHAVIOR REQUIREMENTS"):
            isBehavior = True
            if assistant_message:
                history.append({"role": "assistant", "content": assistant_message})
                assistant_message = ""
                user_message = ""

            continue
        if line.strip().startswith("/*"):
            isUser = True
            if assistant_message:
                if not isBehavior:
                    history.append({"role": "assistant", "content": assistant_message})
                else:
                    requirements_and_code[user_message] = assistant_message
                assistant_message = ""
                user_message = ""
            # user_message = line
        elif line.strip().startswith("*/"):
            isUser = False
            if not isBehavior:
                history.append({"role": "user", "content": user_message})
            
        elif isUser:
            user_message += line
        else:
            assistant_message += line
    if assistant_message:
        if not isBehavior:
            history.append({"role": "assistant", "content": assistant_message})
        else:
            requirements_and_code[user_message] = assistant_message
    
    #history now holds the entity and queries part
    #now we will get the requirements and their corresponding code

    # print (requirements_and_code)
    # print (history)
    #export the history and requirements_and_code to a file called history_and_requirements
    with open("history_and_requirements.txt", "w") as file:
        file.write("History:\n")
        for item in history:
            file.write(f"Role: {item['role']}\n")
            file.write(f"Content: {item['content']}\n")
            file.write("\n")
        
        file.write("Requirements and Code:\n")
        for req, code in requirements_and_code.items():
            file.write(f"Requirement: {req}\n")
            file.write(f"Code: {code}\n")
            file.write("\n")

        
    
    
    # behavior_instructions_type = input("\n\n\Select behavior instructions type: \n1 for Basic, \n2 for BasicPlus, \n3 for DSL, \n3.5 for DSL_Isolated, \n4 for Analysis, \n5 for Default, \n6 for all \n see readme for more information\n your choice:")
    # behavior_instructions_types = []
    # if behavior_instructions_type == "1":
    #     behavior_instructions_types = [BehaviorInstructionType.Basic]
    # elif behavior_instructions_type == "2":
        behavior_instructions_types = [BehaviorInstructionType.BasicPlus]
    # elif behavior_instructions_type == "3":
    #     behavior_instructions_types = [BehaviorInstructionType.DSL]
    # elif behavior_instructions_type == "3.5":
    #     behavior_instructions_types = [BehaviorInstructionType.DSL_Isolated]
    # elif behavior_instructions_type == "4":
    #     behavior_instructions_types = [BehaviorInstructionType.Analysis]
    # elif behavior_instructions_type == "5":
    #     behavior_instructions_types = [BehaviorInstructionType.Default]
    # else:
    #     behavior_instructions_types = [BehaviorInstructionType.Basic, BehaviorInstructionType.BasicPlus, BehaviorInstructionType.DSL, BehaviorInstructionType.DSL_Isolated, BehaviorInstructionType.Analysis]
    behavior_instructions_types = [BehaviorInstructionType.BasicPlus, BehaviorInstructionType.BasicPlusNoExamples]
    behavior_instructions_types = [BehaviorInstructionType.BasicPlusNoExamples]
    
    
    all_models= []
    for behavior_instructions_type in behavior_instructions_types:
        model = MyOpenAIApi(model="gpt-4-turbo-2024-04-09", temp=0)
        behavior_instructions_file_path = behavior_instructions_type.value
        model.history = history
        
        #add first value to be system instructions
        instructions = ""
        with open(behavior_instructions_file_path, "r") as file:
                instructions = file.read()

        model.history.insert(0, {"role": "system", "content": instructions})
        model.History_For_Output = history.copy()
        all_models.append(model)

    requirements_code_generated_blue = {}
    file_name = file_path_of_generated_code.split("/")[-1].split("\\")[-1]
    for model in all_models:
        for req, code in requirements_and_code.items():

            response = model.chat_with_gpt_cumulative(req)
            response_isolated = model.chat_with_gpt_cumulative_isolated_BP_generation(req)
            print("_____________________REQ_start______________________")
            print("_____________ChatGPT___________:\n", response)
            # print("_____________History___________:\n","\n".join([item["content"] for item in model.history[1:]]))
            print("_____________________REQ_end______________________")
            result = calc_codebleu([response], [response_isolated], lang="javascript", weights=(0.1, 0.1, 0.4, 0.4), tokenizer=None)
            requirements_code_generated_blue[req] = {"generated_code": response, "generated_code_isolated": response_isolated, "blue_score": result, "file_name": file_name, "behavior_instructions_type": behavior_instructions_type.name}
            print(result)
    #export the requirements_code_generated_blue to a file called requirements_code_generated_blue
    with open("src/BP_code_generation/Evaluation/requirements_code_generated_blue.txt", "w") as file:
        for req, data in requirements_code_generated_blue.items():
            file.write(f"Requirement: {req}\n")
            file.write(f"Generated Code: {data['generated_code']}\n")
            file.write(f"Generated Code Isolated: {data['generated_code_isolated']}\n")
            file.write(f"Blue Score: {data['blue_score']}\n")
            file.write("\n")

    #export it to an excel file
    #in the next format:
    # | Requirement | Code | Generated Code | Blue Score | File Name | Behavior Instructions Type |
    # Convert the dictionary to a DataFrame
    df = pd.DataFrame.from_dict(requirements_code_generated_blue, orient='index')

    # Reset index to get 'Requirement' as a column
    df.reset_index(inplace=True)

    # Rename columns
    df.columns = ['Requirement', 'Generated Code', 'Generated Code Isolated', 'Blue Score', 'Filename', 'Behavior Instructions Type']

    # Export to Excel
    if not output_file_path:
        output_file_path = "src/BP_code_generation/Evaluation/requirements_code_generated_blue.xlsx"
    #if the file already exists, we will append to it
    if os.path.exists(output_file_path):
        # Load the existing Excel file
        excel_file = output_file_path
        book = load_workbook(excel_file)

        # Load the existing sheet into a pandas DataFrame
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    else:
        df.to_excel(output_file_path, index=False)

    print("Exported to Excel file: ", output_file_path)




    # # return history
def load_data_from_excel(file_path):
    """
    requirements_code_generated_blue[req] = {"code": code, "generated_code": response, "blue_score": result}
    f = pd.DataFrame(requirements_code_generated_blue)
    df.to_excel("requirements_code_generated_blue.xlsx")
    """
    df = pd.read_excel(file_path)
    requirements_code_generated_blue = {}

    #show the properties of the dataframe
    # print(df.columns)
    print(df.head())    


    requirements_code_generated_blue = df.set_index('Requirement').T.to_dict()

        
    #calculate blue score for each requirement
    for req, data in requirements_code_generated_blue.items():
        result = calc_codebleu([data["Code"]], [data["Generated Code"]], lang="javascript", weights=(0.1, 0.1, 0.4, 0.4), tokenizer=None)
        # data["blue_score"] = result
        print(result)
    
def get_avg_blue_score(file_path):
    df = pd.read_excel(file_path)

    # Assume the file has columns 'Filename' and 'Blue Score'
    # Group by 'Filename' and calculate the average BLEU score
    #The Blue Score = {'codebleu': 0.8037232132806461, 'ngram_match_score': 0.4330011686519913, 'weighted_ngram_match_score': 0.45038481030831556, 'syntax_match_score': 0.7884615384615384, 'dataflow_match_score': 1.0}
    #The average BLEU score is the average of the 'codebleu' values
    # Extract the 'codebleu' value from the 'Blue Score' dictionary
    df['Blue Score'] = df['Blue Score'].apply(ast.literal_eval)

    df['codebleu'] = df['Blue Score'].apply(lambda x: x['codebleu'])

    # Group by 'Filename' and calculate the average 'codebleu' score
    average_codebleu_scores = df.groupby('Filename')['codebleu'].mean().reset_index()

    # Rename columns for clarity
    average_codebleu_scores.columns = ['Filename', 'Average_CodeBLEU_Score']


    # Display the results
    print(average_codebleu_scores)
    # Calculate the overall average 'codebleu' score, across all files, but give more weight to files with more requirements
    # The overall average 'codebleu' score is the weighted average of the 'Average_CodeBLEU_Score' values, where the weights are the number of requirements in each file
    # Calculate the total number of requirements in each file
    num_requirements_per_file = df['Filename'].value_counts().reset_index()
    num_requirements_per_file.columns = ['Filename', 'Num_Requirements']
    print("Avarege codebleu score for all files: ", (average_codebleu_scores['Average_CodeBLEU_Score'] * num_requirements_per_file['Num_Requirements']).sum() / num_requirements_per_file['Num_Requirements'].sum())


if __name__ == "__main__":
    # get_data_set_of_file(file_path)
    # load_data_from_excel("requirements_code_generated_blue.xlsx")
    #run the evaluation for all the files in the data folder(use relative path)
    data_folder_path = "src/BP_code_generation/Evaluation/Data"
    output_file_path = "src/BP_code_generation/Evaluation/requirements_code_generated_blue"+time.strftime("%Y%m%d-%H%M%S")+".xlsx"
    # output_file_path = "src/BP_code_generation/Evaluation/requirements_code_generated_blue20240822-000457.xlsx"
    for file_name in os.listdir(data_folder_path):
        file_path = os.path.join(data_folder_path, file_name)
        #if the file is a js file
        if file_name.endswith("Zoo.js"):
            # get_data_set_of_file_and_eval(file_path, output_file_path)
            get_data_set_of_file_and_eval_compare_isolated(file_path, output_file_path)


    get_avg_blue_score(output_file_path)
    # print(history)
    # print(requirements_and_code)
    # print(history)
    # print(requirements


# import pandas as pd
# from openpyxl import load_workbook

# # Sample new data to append
# filename = "new_sample_filename.txt"
# new_data = {
#     "req3": {"code": "code3", "generated_code": "gen_code3", "blue_score": 0.90, "filename": filename},
#     "req4": {"code": "code4", "generated_code": "gen_code4", "blue_score": 0.80, "filename": filename},
# }

# # Convert the new data to a DataFrame
# new_df = pd.DataFrame.from_dict(new_data, orient='index')

# # Reset index to get 'Requirement' as a column
# new_df.reset_index(inplace=True)

# # Rename columns
# new_df.columns = ['Requirement', 'Code', 'Generated Code', 'Blue Score', 'Filename']

# # Load the existing Excel file
# excel_file = 'requirements_code_generated_blue.xlsx'
# book = load_workbook(excel_file)

# # Load the existing sheet into a pandas DataFrame
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     writer.book = book
#     new_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

# print("New data appended to requirements_code_generated_blue.xlsx")
